#keep this up to date if a new collection gets released
collections = {
    "furniture":"0xec4de0a00c694cc7957fb90b9005b24a3f4f8b99",
    "clothes":"0x8c15d753c4336617890ff9e82c88aa047762b867",
    "pets":"0x792df6705032cd3ad8a6aa3b3b7b0a42c0b9759c",
    "addons":"0xacc8b12fd8b08ecea19fb586c0c744f423fc3dd2"
}

#dependencies
import subprocess
import sys
import json
import os

def install_dependecies():
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import requests
    import openpyxl
    from openpyxl.styles import PatternFill

try:    
    import requests
    import openpyxl
    from openpyxl.styles import PatternFill
except:
    install_dependecies()

#more variables needed in the whole file 
nftitems = {}
nftshop = requests.get('https://nft.habbo.com/api/shop/items/').json()
ethprice=requests.get('https://min-api.cryptocompare.com/data/price?fsym=ETH&tsyms=USD').json()
eth = ethprice['USD']

#gets the current price per credit
# Only uses 10, 50, 100, 500 and 1000 for average ppc
tokenamounts = ['Habbo Credit 10', 'Habbo Credit 50', 'Habbo Credit 100', 'Habbo Credit 500', 'Habbo Credit 1000']
total = 0
alllisted = 0
for tokenamount in tokenamounts:
    r = requests.get('https://api.x.immutable.com/v1/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address=0xfbf1c1c09a94fe45ea8cc981c478816963ec958c&sell_token_type=ERC721&status=active&sell_token_name='+tokenamount).json()
    for results in r['result']:
        if tokenamount == results['sell']['data']['properties']['name']:    
            pricequantity = results['buy']['data']['quantity']
            pricedecimal = results['buy']['data']['decimals']
            price = int(pricequantity) / (10 ** int(pricedecimal))
            USDprice = price * eth
            tokens = int(tokenamount.replace('Habbo Credit ', ''))
            ppc = round(USDprice/tokens, 3)
            total = total + ppc
            alllisted = alllisted + 1
            break
total = (total/alllisted)

#pushes the products from the NFT shop into the nftitems dictionary
for nfti in nftshop['items']:
    if nfti['collection'] in nftitems.keys():
        nftitems[nfti['collection']][nfti['name']] = nfti['mintCost']
    else:
        nftitems[nfti['collection']] = {}
        nftitems[nfti['collection']][nfti['name']] = nfti['mintCost']

#unquote if you want a json file with items:
# with open("items.json", "w") as fp:
#     json.dump(nftitems , fp)

#loop for every nft item category
for category in nftitems:
    #basic file building
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "Dollar Price"
    sheet.cell(row = 1, column = 3).value = "Credit Price"
    sheet.cell(row = 1, column = 4).value = "P/L"
    sheet['A1'].fill = PatternFill("solid", start_color="c4c2c2")
    sheet['B1'].fill = PatternFill("solid", start_color="c4c2c2")
    sheet['C1'].fill = PatternFill("solid", start_color="c4c2c2")
    sheet['D1'].fill = PatternFill("solid", start_color="c4c2c2")
    
    sorting = {}

    #getting item prices and pushing them to sorting
    for item in nftitems[category]:
        print(item)
        r = requests.get('https://api.x.immutable.com/v1/orders?auxiliary_fee_percentages=1&auxiliary_fee_recipients=0x12cB8E42c7ec27d30df6Cb8f44aa6445D0e1a78C&buy_token_type=ETH&direction=asc&include_fees=true&order_by=buy_quantity_with_fees&page_size=48&sell_token_address='+collections[category]+'&sell_token_type=ERC721&status=active&sell_token_name='+item).json()
        for results in r['result']:
            if item in results['sell']['data']['properties']['name']:
                pricequantity = results['buy']['data']['quantity']
                pricedecimal = results['buy']['data']['decimals']
                price = int(pricequantity) / (10 ** int(pricedecimal))
                usdprice = price * eth
                pl = (usdprice - (total * nftitems[category][item]))/(total * nftitems[category][item]) * 100
                sorting[item] = {"name":item, "DPrice":round(usdprice,2),"CPrice":nftitems[category][item],"PL":(round(pl,2))}
                break

    #sorting the sorting dictionary on P/L. (sorteddict only returns the sortitem name)
    counter = 2
    sorteddict = sorted(sorting, reverse=True, key=lambda x: (sorting[x]['PL']))
    for sortitem in sorteddict:
        sheet.cell(row = counter, column = 1).value = sorting[sortitem]['name']
        sheet.cell(row = counter, column = 2).value = sorting[sortitem]['DPrice']
        sheet.cell(row = counter, column = 3).value = sorting[sortitem]['CPrice']
        sheet.cell(row = counter, column = 4).value = sorting[sortitem]['PL']
        counter = counter+1


    #resizing the rows. Yes I stole this from stackoverflow.
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value

    #saving the excel file
    wb.save(category+".xlsx") 

openfile = input('Open the files? (Y/N): ')
if openfile.lower() == 'y':
    for category in nftitems:
        os.system('start '+category+'.xlsx')
